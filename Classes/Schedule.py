
from Classes.ClassInSchool import ClassInSchool
from Classes.Group import Group
from Classes.ScheduleCell import ScheduleCell
import copy
import csv
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
import ControllerGUI

TESTING = True
ROWS_PER_GROUP = 3

class Schedule:
    def __init__(self):
        self.scheduleTable = []

    def display_schedule(self):
        for day in range(5):
            day_schedule = self.scheduleTable[day]
            print(f"Day {day + 1}")
            for hour in range(5):
                for cell in day_schedule[hour]:
                    cell.display_info()
                    print(',', end = '')
                print()
        print()

    def create_groups(self, all_classes):
        
        for day in range(5):
            day_schedule = []
            prev_groups = [[], [], [], []]
            for hour in range(5):
                day_schedule.append([])
                present_groups = [[], [], [], []]

                # check attendance
                for school_class in all_classes:
                    class_attendance = school_class.get_attendance(day, hour)
                    if class_attendance > 0:
                        present_groups[school_class.get_grade()].append(school_class)
                
                # update group size and elements
                for i in range(4):
                    expelled_groups = [[], [], [], []]
                    for group in prev_groups[i]:
                        group.set_time(day, hour)
                        
                        for school_class in group.get_list_of_classes():
                            for grade in range(4):
                                present_groups[grade] = [c for c in present_groups[grade] if school_class.get_class_name() != c.get_class_name()]

                        # removing classes from too big subgroups
                        for g in range(4):
                            subgroup = sorted(group.get_subgroup(g), key=lambda school_class: school_class.get_attendance(day, hour), reverse=True)
                            for school_class in subgroup:
                                if group.get_grade_attendance(g) > 25:
                                    group.remove_class(school_class)
                                    present_groups[school_class.get_grade()].append(school_class)
                                else:
                                    break
                        
                        # removing subgroups from too big groups
                        subgroups = group.get_subgroups() # starting with the oldest group
                        expelled_group = Group(day, hour)
                        for subgroup in subgroups:
                            if group.get_attendance() > 25:
                                for school_class in subgroup:
                                    group.remove_class(school_class)
                                    if not expelled_group.add_children(school_class):
                                        expelled_groups[expelled_group.get_youngest_grade()].append(expelled_group)
                                        expelled_group = Group(day, hour)
                            else:
                                break
                        if expelled_group.get_attendance != 0:
                            expelled_groups[expelled_group.get_youngest_grade()].append(expelled_group)
                    prev_groups[i] += expelled_groups[i]

                new_groups = [[], [], [], []]
                
                # creating same-level groups
                for grade in range(4):

                    for group in prev_groups[grade]:
                        present_groups[grade].append(group)
                    
                    present_groups[grade] = sorted(present_groups[grade], key=lambda sc: sc.get_attendance(day, hour), reverse=True)

                    while present_groups[grade]:
                        group = Group(day, hour)
                        while present_groups[grade] and group.get_attendance() + present_groups[grade][0].get_attendance(day, hour) <= 25:
                            group.add_children(present_groups[grade].pop(0))
                        if group.get_attendance() > 0:
                            new_groups[grade].append(group)
                
                # creating cross-level groups
                for grade in range(3): # starting with the oldest
                    younger_grade_groups = new_groups[grade]
                    older_grade_groups = new_groups[grade + 1]

                    for younger_group in younger_grade_groups:
                        for older_group in older_grade_groups[:]:
                            if younger_group.get_attendance() + older_group.get_attendance() <= 25:
                                younger_group.add_children(older_group)
                                older_grade_groups.remove(older_group)
                
                for grade in range(2): # starting with the oldest
                    younger_grade_groups = new_groups[grade]
                    older_grade_groups = new_groups[grade + 2]

                    for younger_group in younger_grade_groups:
                        for older_group in older_grade_groups[:]:
                            if younger_group.get_attendance() + older_group.get_attendance() <= 25:
                                younger_group.add_children(older_group)
                                older_grade_groups.remove(older_group)
                
                younger_grade_groups = new_groups[0]
                older_grade_groups = new_groups[3]

                for younger_group in younger_grade_groups:
                    for older_group in older_grade_groups[:]:
                        if younger_group.get_attendance() + older_group.get_attendance() <= 25:
                            younger_group.add_children(older_group)
                            older_grade_groups.remove(older_group)

                # Remove empty groups
                new_groups = [[group for group in grade_groups if group.get_attendance() > 0] for grade_groups in new_groups]

                for grade in range(4):
                    for group in new_groups[grade]:
                        new_cell = ScheduleCell(day, hour, [school_class for subgroup in group.get_subgroups() for school_class in subgroup], group.get_attendance())
                        day_schedule[hour].append(new_cell)

                prev_groups = copy.deepcopy(new_groups)
            self.scheduleTable.append(day_schedule)
                

    def save_as(self, filepath, filename):
        days = ["Godzina", "Poniedziałek", "Wtorek", "Środa", "Czwartek", "Piątek"]

        full_path = os.path.join(filepath, filename + ".csv")
        counter = 1
        while os.path.exists(full_path):
            full_path = os.path.join(filepath, f"{filename}({counter}).csv")
            counter += 1

        with open(full_path, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(days)
            for hour in range(5):
                max_number_of_groups = 0
                for day in range(5):
                    schedule_cells = self.scheduleTable[day][hour]
                    if max_number_of_groups < len(schedule_cells):
                        max_number_of_groups = len(schedule_cells)
                hour_table =  [["" for _ in range(6)] for _ in range(max_number_of_groups)]
                hour_table[0][0] = hour + 1
                for day in range(5):
                    schedule_cells = self.scheduleTable[day][hour]
                    for i, cell in enumerate(schedule_cells):
                        if not TESTING:
                            hour_table[i][day + 1] = cell.get_info()
                        else:
                            hour_table[i][day + 1] = cell.get_info_with_sizes()
                writer.writerows(hour_table)

    def save_as_xlsx(self, filepath, filename):
        column_names = ["Godzina", "Poniedziałek", "Wtorek", "Środa", "Czwartek", "Piątek"]
        #row_names = ["11:30 - 12:30", "12:30 - 13:30", "13:30 - 14:30", "14:30 - 15:30", "15:30 - 16:30", "16:30 - 17:00"]
        row_names = ControllerGUI.day_hours
        wb = Workbook()
        ws = wb.active
        ws.append(column_names)
        wrap_alignment = Alignment(wrap_text=True)
        table_length = 0
        for hour in range(5):
            max_number_of_groups = 0
            for day in range(5):
                schedule_cells = self.scheduleTable[day][hour]
                number_of_groups = 0
                for cell in schedule_cells:
                    if cell.attendance != 0:
                        number_of_groups += 1
                if max_number_of_groups < number_of_groups:
                    max_number_of_groups = number_of_groups

            hour_table = [["" for _ in range(6)] for _ in range(max_number_of_groups)]
            hour_table[0][0] = row_names[hour]
            for day in range(5):
                schedule_cells = []
                for cell in self.scheduleTable[day][hour]:
                    if cell.attendance != 0:
                        schedule_cells.append(cell)
                for i, cell in enumerate(schedule_cells):
                    if not TESTING:
                        hour_table[i][day + 1] = cell.get_info()
                    else:
                        hour_table[i][day + 1] = cell.get_info_with_sizes()
                

            for row in hour_table:
                ws.append(row)
                ws.append(["" for _ in range(6)])
                ws.append(["" for _ in range(6)])
                table_length += ROWS_PER_GROUP

            # Zastosowanie zawijania tekstu do wszystkich komórek
            for row in ws.iter_rows(min_row=1, max_col=6, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = wrap_alignment
        
        # Scalanie komórek
        current_value = None
        start_row = None

        for row in range(2, table_length + 2):
            cell_value = ws[f'A{row}'].value
            if cell_value is not None and cell_value != '':  # Znaleziono nową wartość
                if current_value is not None and start_row is not None:  # Scal poprzednie komórki
                    end_row = row - 1
                    if end_row > start_row:  # Scal tylko jeśli jest co scalać
                        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                current_value = cell_value
                start_row = row
            elif cell_value is None or cell_value == '':  # Pusta komórka
                continue
        if current_value is not None and start_row is not None:
            end_row = table_length + 1
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        
        # Definiowanie kolumn, dla których chcesz ustawić szerokość
        columns = ['A', 'B', 'C', 'D', 'E', 'F']
        empty_block = False
        start_row = None
        for col in range(1, 6):
            for row in range(2, table_length + 2, ROWS_PER_GROUP):
                cell_value = ws[f'{columns[col]}{row}'].value
                if cell_value is None or cell_value == '':  # Pusta komórka
                    if not empty_block:
                        start_row = row
                        empty_block = True
                else:
                    if empty_block:
                        ws.merge_cells(start_row=start_row, start_column=col+1, end_row=row - 1, end_column=col+1)
                        empty_block = False

        # Przypisanie stylu ramki
        side = Side(style='thin', color='ffffff')
        border_style = Border(left=side, right=side, top=side, bottom=side)
        for i in range(1, table_length + 2):
            for j in range(1, 7):
                cell = ws.cell(row=i, column=j)
                cell.border = border_style
                ws[row][col].font = Font(name='Century Gothic', size=10, bold=True, italic=False, color='656565')

        # Ustawianie szerokości kolumn
        for col in columns:
            ws.column_dimensions[col].width = 20
        
        for row in range(2, table_length + 2, ROWS_PER_GROUP):
            for col in range(1, 6):
                color = 'ffffff' #dadada
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                ws[row][col].fill = fill
                ws[row][col].font = Font(name='Century Gothic', size=12, bold=True, italic=False, color='9b547c')
        

        # Ustawianie wysokości i wyśrodkowanie wierszy na nauczycieli
        for row in range(3, table_length + 3, ROWS_PER_GROUP):
            ws.row_dimensions[row].height = 40
            for col in range(1, 6):
                ws[row][col].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                color = 'eaeaea'
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                ws[row][col].fill = fill
                

        # Wyśrodkowanie wierszy na sale
        for row in range(4, table_length + 4, ROWS_PER_GROUP):
            for col in range(1, 6):
                color = 'eaeaea'
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                ws[row][col].fill = fill
                ws[row][col].alignment = Alignment(horizontal='center', vertical='center')

        # Ustawienie środkowania i kolor dla kolumny A
        for row in range(0, table_length + 1):
            ws['A'][row].alignment = Alignment(horizontal='center', vertical='center', textRotation=90, wrapText = True)
            color = 'cca9bf'
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws['A'][row].fill = fill
            ws['A'][row].font = Font(name='Century Gothic', size=20, bold=True, italic=False, color='ffffff')
            
        # Ustawienie środkowania i kolor dla wiersza 1
        for col in range(0, 6):
            ws.row_dimensions[1].height = 30
            ws[columns[col]][0].alignment = Alignment(horizontal='center', vertical='center')
            color = '656565'
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws[columns[col]][0].fill = fill
            ws[columns[col]][0].font =Font(name='Century Gothic', size=10, bold=True, italic=False, color='ffffff')

        full_path = os.path.join(filepath, filename + ".xlsx")
        if os.path.exists(full_path):
            os.remove(full_path)
        wb.save(full_path)